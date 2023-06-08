import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import openpyxl
import sys

def generate_excel_file_name(prefix='product'):
    folder_path = 'run/bin/temp'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    i = 1
    while True:
        file_name = f"{prefix}_{i}.xlsx"
        file_path = os.path.join(folder_path, file_name)
        if not os.path.isfile(file_path):
            return file_path
        i += 1


def remove_empty_columns(sheet):
    for i in reversed(range(1, sheet.max_column + 1)):
        column = sheet[openpyxl.utils.get_column_letter(i)]
        if all([cell.value is None for cell in column]):
            sheet.delete_cols(i)


# Получение URL-адреса страницы из аргументов командной строки
url = sys.argv[1]

# Создание пустого списка для хранения данных о товарах
data = []

# Запрос на получение HTML-кода страницы
response = requests.get(url)
html_code = response.content

# Парсинг HTML-кода с помощью Beautiful Soup
soup = BeautifulSoup(html_code, 'html.parser')

# Получение списка карточек товаров
products = soup.find_all('div', {'class': 'card'})

# Обход каждой карточки товара и сбор нужной информации
for product in products:
    name = product.find('a', {'class': 'card-title'}).text.strip()
    price = product.find('div', {'class': 'card-price'}).text.strip().replace('\n', '').replace(' ', '')
    data.append({'Name': name, 'Price': price})

# Замена символа табуляции на символ переноса строки
for i in range(len(data)):
    data[i]['Price'] = data[i]['Price'].replace('\t', '\n').strip()

# Создание DataFrame из списка данных
df = pd.DataFrame(data)

# Удаление дубликатов по столбцу 'Name/Price'
df.drop_duplicates(subset=['Name', 'Price'], inplace=True)

# Сортировка по столбцу 'Price' в порядке возрастания цены
df.sort_values(by='Price', ascending=True, inplace=True)

# Генерируем уникальное имя файла
new_file_name = generate_excel_file_name()

# Создаем новый документ Excel с помощью openpyxl
workbook = openpyxl.Workbook()
sheet = workbook.active

# Записываем данные в лист
for index, row in df.iterrows():
    sheet['A'+str(index+1)] = row['Name']
    sheet['B'+str(index+1)] = row['Price']

# Удаление пустых строк
for i in reversed(range(1, sheet.max_row + 1)):
    if all([cell.value is None for cell in sheet[i]]):
        sheet.delete_rows(i)

# Удаляем пустые столбцы
remove_empty_columns(sheet)

# Сохраняем документ с уникальным именем
workbook.save(new_file_name)
print(f"Файл {new_file_name} успешно создан!")
