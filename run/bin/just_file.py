import os
import openpyxl
import datetime

# Define a function to generate a unique file name for the new Excel document
def generate_excel_file_name():
    """
    Generate a unique file name for the new Excel document based on the current timestamp.
    """
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"output_{timestamp}.xlsx"

# Path to the directory where Excel files are located
path_to_excels = './run/bin/temp'

# List of Excel file names in the specified directory
excel_files = [f for f in os.listdir(path_to_excels) if f.endswith('.xlsx')]

# Create a new Excel document
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write data from all files to the new Excel document
for idx, excel_file in enumerate(excel_files):
    # Open the Excel file
    file_path = os.path.join('./run/bin/temp', excel_file)
    wb = openpyxl.load_workbook(file_path)

    # Get the first worksheet
    ws = wb.worksheets[0]

    # Check if column headers exist
    headers_exist = True if (ws.cell(row=1, column=1).value == 'Name' and ws.cell(row=1, column=2).value == 'Price') else False
    
    # Copy column headers from the first file only
    if idx == 0:
        if headers_exist:
            sheet['A1'] = 'Name'
            sheet['B1'] = 'Price'
    
    # Copy data from current worksheet to new document
    for i in range(1 if headers_exist else 1, ws.max_row + 1):
        sheet.append([cell.value for cell in ws[i]])

# Remove empty rows and columns
for i in reversed(range(1, sheet.max_row + 1)):
    if all([cell.value is None for cell in sheet[i]]):
        sheet.delete_rows(i)

for i in reversed(range(1, sheet.max_column + 1)):
    column = sheet[openpyxl.utils.get_column_letter(i)]
    if all([cell.value is None for cell in column]):
        sheet.delete_cols(i)

# Generate a unique file name and save the new Excel document
if not os.path.exists('./save'):
    os.makedirs('./save')

new_file_path = os.path.join('./save', generate_excel_file_name())
workbook.save(new_file_path)
print(f"File {new_file_path} successfully created!")

